import openpyxl
from datetime import timedelta

SAGYOUDETA = input("作業データの日付を入力してください。（４桁数字、例：1021）")
dcmonth = SAGYOUDETA[0:2]
dcday = SAGYOUDETA[2:]

# 工数集計表を開く
try:
    file = openpyxl.load_workbook(r"C:\1\py\【2020{}】ダイエットクック日別勤怠集計表奥山ver.xlsx".format(SAGYOUDETA))
    print("既存データ{}.xlsxを開きました,少々お待ち下さい。".format(SAGYOUDETA))
except FileNotFoundError:
    file = openpyxl.load_workbook(r"C:\1\py\10241517\【20201022】ダイエットクック日別勤怠集計表奥山ver.xlsx")
    print("10241517フォルダ1022雛形を開きました,少々お待ち下さい。")
sheet = file.active

# NEWマスターファイルを開く (10月分)
masterfile = openpyxl.load_workbook(r"C:\1\py\小田原NEWマスタ{}月奥山ver.xlsx".format(dcmonth))
mastersheet = masterfile.active

# NEW*月分勤怠表を開く(10月分)
detailfile = openpyxl.load_workbook(r"C:\1\py\NEW{}月分勤怠表奥山ver.xlsx".format(dcmonth))

# 入力待ち


# 作業日付を入れる
sheet["D1"] = dcday

print("{}月{}日工数入力開始".format(dcmonth, dcday))

while True:

    inputNINZUU = input("人数")
    if inputNINZUU == "":
        break
    inputJIKAN = input("作業時間")

    for iii in range(int(inputNINZUU)):

        userinput = input("番号＋終了時間")
        if len(userinput) == 4:
            userinput = userinput[:2] + inputJIKAN + userinput
        userinput = userinput[:2]+inputJIKAN+userinput
        # no = input("番号を入力してください（４桁数字）。空白値で終了")
        if userinput == "":
            break
        elif len(userinput) != 6 and len(userinput) != 13:
            print("入力に間違いがあります。お確かめの上、もう一度入力してください。")
            continue
        else:
            pass
        # 番号を7000番台に変換
        np7000transfer = int(userinput[0:4]) + 70000000

        if len(userinput) == 6 and userinput[-2:] == "08":
            starttimeH = int(8)
            starttimeM = int(30)
            starttime = timedelta(hours=starttimeH, minutes=starttimeM)

            # 就業時間処理
            endtimeH = int(17)
            endtimeM = int(30)
            endtime = timedelta(hours=endtimeH, minutes=endtimeM)

            # 休憩時間処理
            resttimeH = int(1)
            resttime = timedelta(hours=resttimeH)

        elif len(userinput) == 6 and userinput[-2:] == "10":
            starttimeH = int(10)
            starttimeM = int(00)
            starttime = timedelta(hours=starttimeH, minutes=starttimeM)

            # 就業時間処理
            endtimeH = int(19)
            endtimeM = int(00)
            endtime = timedelta(hours=endtimeH, minutes=endtimeM)

            # 休憩時間処理
            resttimeH = int(1)
            resttime = timedelta(hours=resttimeH)


        elif len(userinput) == 6 and userinput[-2:] == "18":
            starttimeH = int(18)
            starttimeM = int(00)
            starttime = timedelta(hours=starttimeH, minutes=starttimeM)

            # 就業時間処理
            endtimeH = int(24)
            endtimeM = int(00)
            endtime = timedelta(hours=endtimeH, minutes=endtimeM)

            # 休憩時間処理
            resttimeH = int(0)
            resttime = timedelta(hours=resttimeH)

        else:
            # 始業時間処理
            starttimeH = int(userinput[4:6])
            starttimeM = int(userinput[6:8])
            starttime = timedelta(hours=starttimeH, minutes=starttimeM)

            # 就業時間処理
            endtimeH = int(userinput[8:10])
            endtimeM = int(userinput[10:12])
            endtime = timedelta(hours=endtimeH, minutes=endtimeM)

            # 休憩時間処理
            resttimeH = int(userinput[-1])
            resttime = timedelta(hours=resttimeH)

        # 休憩開始、終了時間処理
        # もし出勤時間が8時30分且つ休憩が1時間の場合：
        if starttime >= timedelta(hours=8) and starttime < timedelta(hours=10) and resttime == timedelta(hours=1):
            resttimestart = timedelta(hours=12)
            resttimefinish = timedelta(hours=13)
        # もし出勤時間が10時00分且つ休憩が1時間の場合
        elif starttime >= timedelta(hours=10) and starttime < timedelta(hours=12) and resttime == timedelta(hours=1):
            resttimestart = timedelta(hours=13)
            resttimefinish = timedelta(hours=14)
        # 上記2パターン以外
        elif starttime >= timedelta(hours=18) or starttime == timedelta(hours=19) or resttime == timedelta(hours=0):
            resttimestart = timedelta(hours=0)
            resttimefinish = timedelta(hours=0)
        else:
            print("休憩時間エラー")
            continue

        # 労働時間処理
        worktime = endtime - starttime - resttime

        # print("{}\n{}\n{}\n{}".format(starttime,endtime,resttime,worktime))
        if worktime < timedelta(hours=0):
            print("!!!エラー!!!総労働時間がマイナスになっている。00：00→24：00と入力してください。")

        # 番号特定作業、時間入力
        # B列の番号を探す
        for cell in sheet["B"]:
            # もし入力された番号がヒットしたら
            if cell.value == np7000transfer:
                # まずは名前を表示する
                name = sheet.cell(row=cell.row, column=cell.column + 1).value
                # 確認用
                print("\n\n\n名前【 {} 】\n実働【 {} 】\n\n".format(name, str(worktime)[:-3]))
                # userinput = input("データを入力してください。（9桁数字）例：080019001　←時刻数字開始+終了+最後の数字は休憩時間")
                # starttime = userinput[:4]
                # endtime = userinput[4:9]
                # resttime = userinput[-1]
                # worktime = 0
                # 本番、時間入力
                # try:
                sheet.cell(row=cell.row, column=cell.column + 2).value = starttime
                # except PermissionError:
                #     print("エラー。エクセルを閉じてからこのプログラムを実行してください。")
                #     break
                if endtime == timedelta(hours=24, minutes=00) or endtime == timedelta(hours=00, minutes=00):
                    # もし０時退勤：
                    sheet.cell(row=cell.row, column=cell.column + 5).value = "24:00"
                    sheet.cell(row=cell.row, column=cell.column + 5).number_format = "[h]:mm"
                else:
                    # ０時退勤ではない
                    sheet.cell(row=cell.row, column=cell.column + 5).value = endtime
                # 休憩時間処理
                if resttime == timedelta(hours=00, minutes=00):
                    # もし休憩無し：
                    sheet.cell(row=cell.row, column=cell.column + 8).value = str('')
                else:
                    # 休憩がゼロ以外；
                    sheet.cell(row=cell.row, column=cell.column + 8).value = resttime

                break

        # NEWマスタを入れる。
        for cell in mastersheet["C"]:
            # もし入力された番号がヒットしたら
            if cell.value == np7000transfer:
                # まずは名前を表示する
                name = mastersheet.cell(row=cell.row, column=cell.column + 1).value
                # print("NEWマスタ名前特定：{},{}".format(name, cell))
                todaycell = mastersheet.cell(row=cell.row, column=int(dcday) + 4)
                todaycell.value = worktime
                break

            # print("NEWマスタ名前{}特定できませんでした。確認してください。".format(np7000transfer))

        # NEW*月分勤怠表（個人詳細）にデータ入れる
        # シート名前遍歴
        list = detailfile.sheetnames
        for i in list:
            # もし番号がシート名に含まれている場合
            if str(np7000transfer) in i:
                # シート名
                # print(i)
                detailsheet = detailfile[i]
                # 特定したシートのD4値（確認用）
                # print(detailsheet["D4"].value)
                # 出勤時間を入れる
                detailsheet["C{}".format(int(dcday) + 8)].value = starttime
                # 退勤時間を入れる
                if endtime == timedelta(hours=24, minutes=00):
                    detailsheet["E{}".format(int(dcday) + 8)].value = "24:00"
                    detailsheet["E{}".format(int(dcday) + 8)].number_format = "[h]:mm"
                else:
                    detailsheet["E{}".format(int(dcday) + 8)].value = endtime
                # 休憩開始時間を入れる
                if resttime == timedelta(hours=00, minutes=00):
                    pass
                else:
                    detailsheet["F{}".format(int(dcday) + 8)].value = resttimestart
                    # 休憩終了時間を入れる
                    detailsheet["G{}".format(int(dcday) + 8)].value = resttimefinish
                break
        # print("{}のシートが見つかりません。確認してください。".format(i))
    # detailsheet = detailfile.active

    # print("保存済\n")

    # NEW*月分勤怠表を入れる
print("ファイルを保存しています・・・")
file.save("【2020{}】ダイエットクック日別勤怠集計表奥山ver.xlsx".format(SAGYOUDETA))
masterfile.save("小田原NEWマスタ{}月奥山ver.xlsx".format(dcmonth))
detailfile.save("NEW{}月分勤怠表奥山ver.xlsx".format(dcmonth))

file.close()
masterfile.close()
detailfile.close()
print("保存完了、プログラム終了")
